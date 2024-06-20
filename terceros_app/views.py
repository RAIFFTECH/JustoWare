import csv
from openpyxl import Workbook
from django.views import View
from django.http import HttpResponse
from django.db.models import Q
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Table, TableStyle, ListFlowable, ListItem
from django import forms
from .forms import CrearForm
from .models import TERCEROS

# Para obtener todos los registros
class Lista(LoginRequiredMixin, ListView):
    model = TERCEROS
    form = CrearForm
    template_name = 'lista_terceros.html'

# Para buscar por nombre del tercero
def Buscar(request):
    resultados = None
    if 'q' in request.GET:
        query = request.GET['q']
        resultados = TERCEROS.objects.filter(Q(nombre__icontains=query) | Q(doc_ide__icontains=query) | Q(celular1__icontains=query) | Q(email__icontains=query))
    return render(request, 'lista_terceros.html', {'resultados': resultados})

# Para obtener todos los detalles de un registro
class Detalles(LoginRequiredMixin, DetailView):
    model = TERCEROS
    form = CrearForm
    template_name = 'detalles_tercero.html'

# Para crear un nuevo registro
class Crear(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = TERCEROS
    form = CrearForm
    fields = "__all__"
    # fields = ['oficina', 'per_con', 'codigo', 'nom_cto', 'nombre','doc_admin', 'doc_caja', 'inicio_nuevo_per', 'consecutivo']
    template_name = 'crear_tercero.html'

    # Mensaje que se mostrará cuando se inserte el registro
    success_message = 'Registro añadido correctamente.'

    # Redirigimos a la página principal tras insertar el registro
    def get_success_url(self):
        return reverse('listar_terceros')

# Para modificar un registro
class Actualizar(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = TERCEROS
    form = CrearForm
    fields = "__all__"
    # fields = ['oficina', 'per_con', 'codigo', 'nom_cto', 'nombre','doc_admin', 'doc_caja', 'inicio_nuevo_per', 'consecutivo']
    template_name = 'actualizar_tercero.html'
    # Mensaje que se mostrará cuando se actualice el registro
    success_message = 'Registro actualizado correctamente.'

    # Redireccionamos a la página principal tras actualizar el registro
    def get_success_url(self):
        return reverse('listar_terceros')

# Para eliminar un registro
class Eliminar(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = TERCEROS
    form = CrearForm
    fields = "__all__"

    # Redireccionamos a la página principal tras de eliminar el registro
    def get_success_url(self):
        # Mensaje que se mostrará cuando se elimine el registro
        success_message = 'Registro eliminado correctamente.'
        messages.success(self.request, (success_message))
        return reverse('listar_terceros')

# Para imprimir los registros
class ImprimirPDF(View):
    def get(self, request, *args, **kwargs):
        
        terceros = TERCEROS.objects.all()

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="terceros.pdf"'

        # Creamos un objeto PDF con ReportLab
        p = canvas.Canvas(response, pagesize=letter)

        # Creamos una lista para almacenar los elementos de la lista
        elementos_lista = []

        # Llenamos la lista con datos del modelo
        for dato in terceros:
            elemento = ListItem("Tipo Documento: {}, Número Documento: {}, DV: {}, Nit Rápido: {}, Ciudad Expedición Documento: {}, Ciudad Residencia: {}, Tipo Régimen: {}, Fecha Expedición Documento: {}, Tipo Tercero: {}, Primer Apellido: {}, Segundo Apellido: {}, Primer Nombre: {}, Segundo Nombre: {}, Razón Social: {}, Dirección: {}, Código Postal: {}, Teléfono Oficina: {}, Teléfono Residencia: {}, Celular 1: {}, Celular 2: {}, Fax: {}, e-mail: {}, Nombre: {}, Fecha Actualización: {}, Observación: {}, Persona Expuesta PEP: {}, Nit Interno: {}".format(dato.cla_doc, dato.doc_ide, dato.dig_ver, dato.nit_rap, dato.cod_ciu_exp, dato.cod_ciu_res, dato.regimen, dato.fec_exp_ced, dato.tip_ter, dato.pri_ape, dato.seg_ape, dato.pri_nom, dato.seg_nom, dato.raz_soc, dato.direccion, dato.cod_pos, dato.tel_ofi, dato.tel_res, dato.celular1, dato.celular2, dato.fax, dato.email, dato.nombre, dato.fec_act, dato.observacion, dato.per_pub_exp, dato.nit_interno))
            elementos_lista.append(elemento)

        # Creamos la lista con columnas
        lista_columnas = ListFlowable(
            elementos_lista, bulletType='bullet', start='circle')

        # Dibujamos la lista en el PDF
        lista_columnas.wrapOn(p, 100, 100)
        lista_columnas.drawOn(p, 100, 600)
        
        
        
        #  # Creamos un objeto PDF con ReportLab
        # p = canvas.Canvas(response, pagesize=letter)

        # # Creamos una lista para almacenar los datos de la tabla
        # datos_tabla = [["Tipo Documento: {}, Número Documento: {}, DV: {}, Nit Rápido: {}, Ciudad Expedición Documento: {}, Ciudad Residencia: {}, Tipo Régimen: {}, Fecha Expedición Documento: {}, Tipo Tercero: {}, Primer Apellido: {}, Segundo Apellido: {}, Primer Nombre: {}, Segundo Nombre: {}, Razón Social: {}, Dirección: {}, Código Postal: {}, Teléfono Oficina: {}, Teléfono Residencia: {}, Celular 1: {}, Celular 2: {}, Fax: {}, e-mail: {}, Nombre: {}, Fecha Actualización: {}, Observación: {}, Persona Expuesta PEP: {}, Nit Interno: {}"

        # # Llenamos la lista con datos del modelo
        # for dato in terceros:
        #     datos_tabla.append([dato.cla_doc, dato.doc_ide, dato.dig_ver, dato.nit_rap, dato.cod_ciu_exp, dato.cod_ciu_res, dato.regimen, dato.fec_exp_ced, dato.tip_ter, dato.pri_ape, dato.seg_ape, dato.pri_nom, dato.seg_nom, dato.raz_soc, dato.direccion, dato.cod_pos, dato.tel_ofi, dato.tel_res, dato.celular1, dato.celular2, dato.fax, dato.email, dato.nombre, dato.fec_act, dato.observacion, dato.per_pub_exp, dato.nit_interno])
        # #     y -=20


        # # Creamos la tabla
        # tabla = Table(datos_tabla)

        # # Aplicamos estilos a la tabla
        # estilo_tabla = TableStyle([('BACKGROUND', (0, 0), (-1, 0), '#77a4d4'),
        #                           ('TEXTCOLOR', (0, 0), (-1, 0), (255, 255, 255)),
        #                           ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        #                           ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        #                           ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        #                           ('BACKGROUND', (0, 1), (-1, -1), '#f7f7f7'),
        #                           ('GRID', (0, 0), (-1, -1), 1, '#cccccc')])

        # tabla.setStyle(estilo_tabla)

        # # Dibujamos la tabla en el PDF
        # tabla.wrapOn(p, 0, 0)
        # tabla.drawOn(p, 30, 600)


        # # Creamos un objeto PDF con ReportLab
        # p = canvas.Canvas(response)

        # # Agregamos contenido al PDF utilizando datos de la base de datos
        # y = 800
        # for dato in terceros:

        #     # p.drawString(60, y, f"Cliente: {dato.cliente}")
        #     p.drawString(60, y, f"Tipo Documento: {dato.cla_doc}, Número Documento: {dato.doc_ide}, DV: {dato.dig_ver}, Nit Rápido: {dato.nit_rap}, Ciudad Expedición Documento: {dato.cod_ciu_exp}, Ciudad Residencia: {dato.cod_ciu_res}, Tipo Régimen: {dato.regimen}, Fecha Expedición Documento: {dato.fec_exp_ced}, Tipo Tercero: {dato.tip_ter}, Primer Apellido: {dato.pri_ape}, Segundo Apellido: {dato.seg_ape}, Primer Nombre: {dato.pri_nom}, Segundo Nombre: {dato.seg_nom}, Razón Social: {dato.raz_soc}, Dirección: {dato.direccion}, Código Postal: {dato.cod_pos}, Teléfono Oficina: {dato.tel_ofi}, Teléfono Residencia: {dato.tel_res}, Celular 1: {dato.celular1}, Celular 2: {dato.celular2}, Fax: {dato.fax}, e-mail: {dato.email}, Nombre: {dato.nombre}, Fecha Actualización: {dato.fec_act}, Observación: {dato.observacion}, Persona Expuesta PEP: {dato.per_pub_exp}, Nit Interno: {dato.nit_interno}")
        #     y -= 20


            # Agrega más campos según tus necesidades
            # p.drawString(60, 800, f"Cliente: {dato.cliente}")
            # p.drawString(60, 780, f"Tipo Documento: {dato.cla_doc}")
            # p.drawString(60, 760, f"Número Documento: {dato.doc_ide}")
            # p.drawString(60, 740, f"DV: {dato.dig_ver}")
            # p.drawString(60, 720, f"Nit Rápido: {dato.nit_rap}")
            # p.drawString(60, 700, f"Ciudad Expedición Documento: {dato.cod_ciu_exp}")
            # p.drawString(60, 680, f"Ciudad Residencia: {dato.cod_ciu_res}")
            # p.drawString(60, 660, f"Tipo Régimen: {dato.regimen}")
            # p.drawString(60, 640, f"Fecha Expedición Documento: {dato.fec_exp_ced}")
            # p.drawString(60, 620, f"Tipo Tercero: {dato.tip_ter}")
            # p.drawString(60, 600, f"Primer Apellido: {dato.pri_ape}")
            # p.drawString(60, 580, f"Segundo Apellido: {dato.seg_ape}")
            # p.drawString(60, 560, f"Primer Nombre: {dato.pri_nom}")
            # p.drawString(60, 540, f"Segundo Nombre: {dato.seg_nom}")
            # p.drawString(60, 520, f"Razón Social: {dato.raz_soc}")
            # p.drawString(60, 500, f"Dirección: {dato.direccion}")
            # p.drawString(60, 480, f"Código Postal: {dato.cod_pos}")
            # p.drawString(60, 460, f"Teléfono Oficina: {dato.tel_ofi}")
            # p.drawString(60, 440, f"Teléfono Residencia: {dato.tel_res}")
            # p.drawString(60, 400, f"Celular 1: {dato.celular1}")
            # p.drawString(60, 380, f"Celular 2: {dato.celular2}")
            # p.drawString(60, 360, f"Fax: {dato.fax}")
            # p.drawString(60, 340, f"e-mail: {dato.email}")
            # p.drawString(60, 320, f"Nombre: {dato.nombre}")
            # p.drawString(60, 300, f"Fecha Actualización: {dato.fec_act}")
            # p.drawString(60, 280, f"Observación: {dato.observacion}")
            # p.drawString(60, 260, f"Persona Expuesta PEP: {dato.per_pub_exp}")
            # p.drawString(60, 240, f"Nit Interno: {dato.nit_interno}")

            # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response

# Para imprimir un registro
class ImprimePDF(View):
    def get(self, request, pk, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        dato = TERCEROS.objects.get(pk=pk)

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="terceros.pdf"'

        # Creamos un objeto PDF con ReportLab
        p = canvas.Canvas(response)

        p.drawString(60, 800, f"Cliente: {dato.cliente}")
        p.drawString(60, 780, f"Tipo Documento: {dato.cla_doc}")
        p.drawString(60, 760, f"Número Documento: {dato.doc_ide}")
        p.drawString(60, 740, f"DV: {dato.dig_ver}")
        p.drawString(60, 720, f"Nit Rápido: {dato.nit_rap}")
        p.drawString(60, 700, f"Ciudad Expedición Documento: {dato.cod_ciu_exp}")
        p.drawString(60, 680, f"Ciudad Residencia: {dato.cod_ciu_res}")
        p.drawString(60, 660, f"Tipo Régimen: {dato.regimen}")
        p.drawString(60, 640, f"Fecha Expedición Documento: {dato.fec_exp_ced}")
        p.drawString(60, 620, f"Tipo Tercero: {dato.tip_ter}")
        p.drawString(60, 600, f"Primer Apellido: {dato.pri_ape}")
        p.drawString(60, 580, f"Segundo Apellido: {dato.seg_ape}")
        p.drawString(60, 560, f"Primer Nombre: {dato.pri_nom}")
        p.drawString(60, 540, f"Segundo Nombre: {dato.seg_nom}")
        p.drawString(60, 520, f"Razón Social: {dato.raz_soc}")
        p.drawString(60, 500, f"Dirección: {dato.direccion}")
        p.drawString(60, 480, f"Código Postal: {dato.cod_pos}")
        p.drawString(60, 460, f"Teléfono Oficina: {dato.tel_ofi}")
        p.drawString(60, 440, f"Teléfono Residencia: {dato.tel_res}")
        p.drawString(60, 400, f"Celular 1: {dato.celular1}")
        p.drawString(60, 380, f"Celular 2: {dato.celular2}")
        p.drawString(60, 360, f"Fax: {dato.fax}")
        p.drawString(60, 340, f"e-mail: {dato.email}")
        p.drawString(60, 320, f"Nombre: {dato.nombre}")
        p.drawString(60, 300, f"Fecha Actualización: {dato.fec_act}")
        p.drawString(60, 280, f"Observación: {dato.observacion}")
        p.drawString(60, 260, f"Persona Expuesta PEP: {dato.per_pub_exp}")
        p.drawString(60, 240, f"Nit Interno: {dato.nit_interno}")
        # Agrega más campos según tus necesidades

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response


class ExportarView(View):
    template_name = 'exportar.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        export_type = request.POST.get('export_type')

        if export_type == 'pdf':
            return self.exportar_pdf()
        elif export_type == 'excel':
            return self.exportar_excel()
        elif export_type == 'csv':
            return self.exportar_csv()

    def exportar_pdf(self):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="exportacion.pdf"'
        terceros = TERCEROS.objects.all()

        p = canvas.Canvas(response)

        for dato in terceros:
            p.drawString(60, 800, f"Cliente: {dato.cliente}")
            p.drawString(60, 780, f"Tipo Documento: {dato.cla_doc}")
            p.drawString(60, 760, f"Número Documento: {dato.doc_ide}")
            p.drawString(60, 740, f"DV: {dato.dig_ver}")
            p.drawString(60, 720, f"Nit Rápido: {dato.nit_rap}")
            p.drawString(60, 700, f"Ciudad Expedición Documento: {dato.cod_ciu_exp}")
            p.drawString(60, 680, f"Ciudad Residencia: {dato.cod_ciu_res}")
            p.drawString(60, 660, f"Tipo Régimen: {dato.regimen}")
            p.drawString(60, 640, f"Fecha Expedición Documento: {dato.fec_exp_ced}")
            p.drawString(60, 620, f"Tipo Tercero: {dato.tip_ter}")
            p.drawString(60, 600, f"Primer Apellido: {dato.pri_ape}")
            p.drawString(60, 580, f"Segundo Apellido: {dato.seg_ape}")
            p.drawString(60, 560, f"Primer Nombre: {dato.pri_nom}")
            p.drawString(60, 540, f"Segundo Nombre: {dato.seg_nom}")
            p.drawString(60, 520, f"Razón Social: {dato.raz_soc}")
            p.drawString(60, 500, f"Dirección: {dato.direccion}")
            p.drawString(60, 480, f"Código Postal: {dato.cod_pos}")
            p.drawString(60, 460, f"Teléfono Oficina: {dato.tel_ofi}")
            p.drawString(60, 440, f"Teléfono Residencia: {dato.tel_res}")
            p.drawString(60, 400, f"Celular 1: {dato.celular1}")
            p.drawString(60, 380, f"Celular 2: {dato.celular2}")
            p.drawString(60, 360, f"Fax: {dato.fax}")
            p.drawString(60, 340, f"e-mail: {dato.email}")
            p.drawString(60, 320, f"Nombre: {dato.nombre}")
            p.drawString(60, 300, f"Fecha Actualización: {dato.fec_act}")
            p.drawString(60, 280, f"Observación: {dato.observacion}")
            p.drawString(60, 260, f"Persona Expuesta PEP: {dato.per_pub_exp}")
            p.drawString(60, 240, f"Nit Interno: {dato.nit_interno}")
            # Agrega más campos según tus necesidades

            # Agrega un salto de página para el siguiente conjunto de datos
            p.showPage()
        # Puedes agregar más contenido según tus necesidades.
        p.save()

        return response

    def exportar_excel(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportacion.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        # Añade encabezados a la hoja de cálculo utilizando los nombres de campo del modelo
        headers = [field.name for field in TERCEROS._meta.fields]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Añade datos a la hoja de cálculo
        terceros = TERCEROS.objects.all()
        for row_num, data in enumerate(terceros, start=2):
            for col_num, field in enumerate(headers, 1):
                sheet.cell(row=row_num, column=col_num,
                           value=getattr(data, field))

        workbook.save(response)
        return response

    def exportar_csv(self):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="exportacion.csv"'

        writer = csv.writer(response)
        # Añade más encabezados según tus necesidades
        writer.writerow(['ID', 'Nombre'])

        terceros = TERCEROS.objects.all()
        for data in terceros:
            # Añade más campos según tus necesidades
            writer.writerow([data.pk, data.nombre])

        return response
