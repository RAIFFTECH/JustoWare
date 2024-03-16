import csv
from openpyxl import Workbook
from django.views import View
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

from django import forms
from .forms import CrearForm
from .models import PLAN_CTAS

# Para obtener todos los registros
class Lista(LoginRequiredMixin, ListView):
    model = PLAN_CTAS
    form = CrearForm
    template_name = 'lista_cuentas.html'
    ordering = ['cliente','per_con','cod_cta']


# Para obtener todos los detalles de un registro
class Detalles(LoginRequiredMixin, DetailView):
    model = PLAN_CTAS
    form = CrearForm
    template_name = 'detalles_cuenta.html'

# Para crear un nuevo registro
class Crear(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = PLAN_CTAS
    form = CrearForm
    # fields = '__all__'
    fields = ['cliente','per_con','cod_cta','nom_cta','tip_cta','dinamica','naturaleza','activa','por_tercero','cta_act_fij','cta_pre','cta_bal','cta_res','cta_ord','cta_ban','cta_gan_per','cta_per_gan','cta_dep','cta_ing_ret','cta_ret_iva','cta_rec']
    template_name = 'crear_cuenta.html'

    # Mensaje que se mostrará cuando se inserte el registro
    success_message = 'Registro añadido correctamente.'

    # Redirigimos a la página principal tras insertar el registro
    def get_success_url(self):
        return reverse('listar_cuenta')

# Para modificar un registro
class Actualizar(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = PLAN_CTAS
    form = CrearForm
    # fields = '__all__'
    fields = ['cliente', 'per_con', 'cod_cta', 'nom_cta', 'tip_cta', 'dinamica', 'naturaleza', 'activa', 'por_tercero', 'cta_act_fij','cta_pre', 'cta_bal', 'cta_res', 'cta_ord', 'cta_ban', 'cta_gan_per', 'cta_per_gan', 'cta_dep', 'cta_ing_ret', 'cta_ret_iva', 'cta_rec']
    template_name = 'actualizar_cuenta.html'
    # Mensaje que se mostrará cuando se actualice el registro
    success_message = 'Registro actualizado correctamente.'

    # Redireccionamos a la página principal tras actualizar el registro
    def get_success_url(self):
        return reverse('listar_cuenta')

# Para eliminar un registro
class Eliminar(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = PLAN_CTAS
    form = CrearForm
    fields = "__all__"

    # Redireccionamos a la página principal tras de eliminar el registro
    def get_success_url(self):
        # Mensaje que se mostrará cuando se elimine el registro
        success_message = 'Registro eliminado correctamente.'
        messages.success(self.request, (success_message))
        return reverse('listar_cuenta')

# Para imprimir los registros
class ImprimirPDF(View):
    def get(self, request, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        cuentas = PLAN_CTAS.objects.all().order_by('cliente','per_con','cod_cta')

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="cuentas.pdf"'

        # Creamos un objeto PDF con ReportLab
        # p = canvas.Canvas(response)
        p = canvas.Canvas(response, pagesize=letter)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        # for dato in cuentas:
        #     p.drawString(80, 800, f"Cliente: {dato.cliente}")
        #     p.drawString(80, 780, f"Periodo Contable: {dato.per_con}")
        #     p.drawString(80, 760, f"Código Cuenta: {dato.cod_cta}")
        #     p.drawString(80, 740, f"Nombre Cuenta: {dato.nom_cta}")
        #     p.drawString(80, 720, f"Tipo Cuenta: {dato.tip_cta}")
        #     p.drawString(80, 700, f"Dinámica: {dato.dinamica}")
        #     p.drawString(80, 680, f"Naturaleza: {dato.naturaleza}")
        #     p.drawString(80, 660, f"Cuenta Activa?: {dato.activa}")
        #     p.drawString(80, 640, f"Contabiliza por Tercero?: {dato.por_tercero}")
        #     p.drawString(80, 620, f"Cuenta Activo Fijo?: {dato.cta_act_fij}")
        #     p.drawString(80, 600, f"Cuenta Presupuesto?: {dato.cta_pre}")
        #     p.drawString(80, 580, f"Cuenta de Balance?: {dato.cta_bal}")
        #     p.drawString(80, 560, f"Cuenta de Resultados?: {dato.cta_res}")
        #     p.drawString(80, 540, f"Cuenta de Orden?: {dato.cta_ord}")
        #     p.drawString(80, 520, f"Cuenta de Banco?: {dato.cta_ban}")
        #     p.drawString(80, 500, f"Cuenta Ganancias?: {dato.cta_gan_per}")
        #     p.drawString(80, 480, f"Cuenta Pérdidas?: {dato.cta_per_gan}")
        #     p.drawString(80, 460, f"Cuenta Depreciación?: {dato.cta_dep}")
        #     p.drawString(80, 440, f"Cuenta Ingresos y Retenciones?: {dato.cta_ing_ret}")
        #     p.drawString(80, 420, f"Cuenta Reteiva?: {dato.cta_ret_iva}")
        #     p.drawString(80, 400, f"Cuenta Recíproca?: {dato.cta_rec}")

        datos_tabla = [["Periodo Contable", "Código Cuenta","Nombre Cuenta","Tipo Cuenta","Dinámica","Naturaleza","Cuenta Activa?","Contabiliza por Tercero?","Cuenta Activo Fijo?","Cuenta Presupuesto?","Cuenta de Balance?","Cuenta de Resultados?","Cuenta de Orden?","Cuenta de Banco?","Cuenta Ganancias?","Cuenta Pérdidas?","Cuenta Depreciación?","Cuenta Ingresos y Retenciones?","Cuenta Reteiva?","Cuenta Recíproca?"]]

        for dato in cuentas:
            datos_tabla.append([dato.per_con,dato.cod_cta,dato.nom_cta,dato.tip_cta,dato.dinamica,dato.naturaleza,dato.activa,dato.por_tercero,dato.cta_act_fij,dato.cta_pre,dato.cta_bal,dato.cta_res,dato.cta_ord,dato.cta_ban,
            dato.cta_gan_per,dato.cta_per_gan,dato.cta_dep,dato.cta_ing_ret,dato.cta_ret_iva,dato.cta_rec])
 
            # Agrega más campos según tus necesidades

        tabla = Table(datos_tabla)

        # estilo_tabla = TableStyle([('BACKGROUND', (0, 0), (-1, 0), '#06153C'),
        #                            ('TEXTCOLOR', (0, 0), (-1, 0), (255, 255, 255)),
        #                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        #                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        #                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        #                            ('BACKGROUND', (0, 1), (-1, -1), '#f7f7f7'),
        #                            ('GRID', (0, 0), (-1, -1), 1, '#06153C')])

        # tabla.setStyle(estilo_tabla)

        # Dibujamos la tabla en el PDF
        tabla.wrapOn(p, 0, 0)
        tabla.drawOn(p, 30, 600)

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response

# Para imprimir un registro
class ImprimePDF(View):
    def get(self, request, pk, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        dato = PLAN_CTAS.objects.get(pk=pk)

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="cuentas.pdf"'

        # Creamos un objeto PDF con ReportLab
        p = canvas.Canvas(response)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        p.drawString(80, 800, f"Cliente: {dato.cliente}")
        p.drawString(80, 780, f"Periodo Contable: {dato.per_con}")
        p.drawString(80, 760, f"Código Cuenta: {dato.cod_cta}")
        p.drawString(80, 740, f"Nombre Cuenta: {dato.nom_cta}")
        p.drawString(80, 720, f"Tipo Cuenta: {dato.tip_cta}")
        p.drawString(80, 700, f"Dinámica: {dato.dinamica}")
        p.drawString(80, 680, f"Naturaleza: {dato.naturaleza}")
        p.drawString(80, 660, f"Cuenta Activa?: {dato.activa}")
        p.drawString(80, 640, f"Contabiliza por Tercero?: {dato.por_tercero}")
        p.drawString(80, 620, f"Cuenta Activo Fijo?: {dato.cta_act_fij}")
        p.drawString(80, 600, f"Cuenta Presupuesto?: {dato.cta_pre}")
        p.drawString(80, 580, f"Cuenta de Balance?: {dato.cta_bal}")
        p.drawString(80, 560, f"Cuenta de Resultados?: {dato.cta_res}")
        p.drawString(80, 540, f"Cuenta de Orden?: {dato.cta_ord}")
        p.drawString(80, 520, f"Cuenta de Banco?: {dato.cta_ban}")
        p.drawString(80, 500, f"Cuenta Ganancias?: {dato.cta_gan_per}")
        p.drawString(80, 480, f"Cuenta Pérdidas?: {dato.cta_per_gan}")
        p.drawString(80, 460, f"Cuenta Depreciación?: {dato.cta_dep}")
        p.drawString(80, 440, f"Cuenta Ingresos y Retenciones?: {dato.cta_ing_ret}")
        p.drawString(80, 420, f"Cuenta Reteiva?: {dato.cta_ret_iva}")
        p.drawString(80, 400, f"Cuenta Recíproca?: {dato.cta_rec}")

        # Agrega más campos según tus necesidades

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response


class ExportarView(View):
    template_name = 'exportar.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        export_type = request.POST.get('export_type')

        if export_type == 'pdf':
            return self.exportar_pdf()
        elif export_type == 'excel':
            return self.exportar_excel()
        elif export_type == 'csv':
            return self.exportar_csv()

    def exportar_pdf(self):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="exportacion.pdf"'
        cuentas = PLAN_CTAS.objects.all()

        p = canvas.Canvas(response)

        for dato in cuentas:
            p.drawString(80, 800, f"Cliente: {dato.cliente}")
            p.drawString(80, 780, f"Periodo Contable: {dato.per_con}")
            p.drawString(80, 760, f"Código Cuenta: {dato.cod_cta}")
            p.drawString(80, 740, f"Nombre Cuenta: {dato.nom_cta}")
            p.drawString(80, 720, f"Tipo Cuenta: {dato.tip_cta}")
            p.drawString(80, 700, f"Dinámica: {dato.dinamica}")
            p.drawString(80, 680, f"Naturaleza: {dato.naturaleza}")
            p.drawString(80, 660, f"Cuenta Activa?: {dato.activa}")
            p.drawString(80, 640, f"Contabiliza por Tercero?: {dato.por_tercero}")
            p.drawString(80, 620, f"Cuenta Activo Fijo?: {dato.cta_act_fij}")
            p.drawString(80, 600, f"Cuenta Presupuesto?: {dato.cta_pre}")
            p.drawString(80, 580, f"Cuenta de Balance?: {dato.cta_bal}")
            p.drawString(80, 560, f"Cuenta de Resultados?: {dato.cta_res}")
            p.drawString(80, 540, f"Cuenta de Orden?: {dato.cta_ord}")
            p.drawString(80, 520, f"Cuenta de Banco?: {dato.cta_ban}")
            p.drawString(80, 500, f"Cuenta Ganancias?: {dato.cta_gan_per}")
            p.drawString(80, 480, f"Cuenta Pérdidas?: {dato.cta_per_gan}")
            p.drawString(80, 460, f"Cuenta Depreciación?: {dato.cta_dep}")
            p.drawString(80, 440, f"Cuenta Ingresos y Retenciones?: {dato.cta_ing_ret}")
            p.drawString(80, 420, f"Cuenta Reteiva?: {dato.cta_ret_iva}")
            p.drawString(80, 400, f"Cuenta Recíproca?: {dato.cta_rec}")

            # Agrega más campos según tus necesidades

            # Agrega un salto de página para el siguiente conjunto de datos
            p.showPage()
        # Puedes agregar más contenido según tus necesidades.
        p.save()

        return response

    def exportar_excel(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportacion.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        # Añade encabezados a la hoja de cálculo utilizando los nombres de campo del modelo
        headers = [field.name for field in PLAN_CTAS._meta.fields]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Añade datos a la hoja de cálculo
        cuentas = PLAN_CTAS.objects.all()
        for row_num, data in enumerate(cuentas, start=2):
            for col_num, field in enumerate(headers, 1):
                sheet.cell(row=row_num, column=col_num,
                           value=getattr(data, field))

        workbook.save(response)
        return response

    def exportar_csv(self):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="exportacion.csv"'

        writer = csv.writer(response)
        # Añade más encabezados según tus necesidades
        writer.writerow(['ID', 'Nombre'])

        cuentas = PLAN_CTAS.objects.all()
        for data in cuentas:
            # Añade más campos según tus necesidades
            writer.writerow([data.pk, data.nombre])

        return response
