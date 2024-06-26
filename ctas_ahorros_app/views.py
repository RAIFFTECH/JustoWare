import csv
from openpyxl import Workbook
from django.views import View
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from reportlab.pdfgen import canvas
from django import forms
from .forms import CrearForm
from .models import CTAS_AHORRO

# Para obtener todos los registros
class Lista(LoginRequiredMixin, ListView):
    model = CTAS_AHORRO
    form = CrearForm
    template_name = 'lista_cta_aho.html'

# Para obtener todos los detalles de un registro
class Detalles(LoginRequiredMixin, DetailView):
    model = CTAS_AHORRO
    form = CrearForm
    template_name = 'detalles_cta_aho.html'

# Para crear un nuevo registro
class Crear(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = CTAS_AHORRO
    form = CrearForm
    fields = '__all__'
    template_name = 'crear_cta_aho.html'

    # Mensaje que se mostrará cuando se inserte el registro
    success_message = 'Registro añadido correctamente.'

    # Redirigimos a la página principal tras insertar el registro
    def get_success_url(self):
        return reverse('listar_cta_aho')

# Para modificar un registro
class Actualizar(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = CTAS_AHORRO
    form = CrearForm
    fields = '__all__'
    template_name = 'actualizar_cta_aho.html'
    # Mensaje que se mostrará cuando se actualice el registro
    success_message = 'Registro actualizado correctamente.'

    # Redireccionamos a la página principal tras actualizar el registro
    def get_success_url(self):
        return reverse('listar_cta_aho')

# Para eliminar un registro
class Eliminar(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = CTAS_AHORRO
    form = CrearForm
    fields = "__all__"

    # Redireccionamos a la página principal tras de eliminar el registro
    def get_success_url(self):
        # Mensaje que se mostrará cuando se elimine el registro
        success_message = 'Registro eliminado correctamente.'
        messages.success(self.request, (success_message))
        return reverse('listar_cta_aho')

# Para imprimir los registros
class ImprimirPDF(View):
    def get(self, request, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        cta_ahos = CTAS_AHORRO.objects.all()

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="cta_ahos.pdf"'

        # Creamos un objeto PDF con ReportLab
        p = canvas.Canvas(response)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        for dato in cta_ahos:
            p.drawString(80, 800, f"Oficina: {dato.oficina}")
            p.drawString(80, 780, f"Línea de Ahorro: {dato.lin_aho}")
            p.drawString(80, 760, f"Nombre Asociado: {dato.asociado}")
            p.drawString(80, 740, f"Número Cuenta: {dato.num_cta}")
            p.drawString(80, 720, f"Estado Cuenta: {dato.est_cta}")
            p.drawString(80, 700, f"Fecha Apertura: {dato.fec_apertura}")
            p.drawString(80, 680, f"Fecha Cancelación: {dato.fec_cancela}")
            p.drawString(80, 660, f"Exenta 4x1000?: {dato.exc_tas_mil}")
            p.drawString(80, 640, f"Fecha Exención: {dato.fec_ini_exc}")
            
            # Agrega más campos según tus necesidades

            # Agrega un salto de página para el siguiente conjunto de datos
            p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response

# Para imprimir un registro
class ImprimePDF(View):
    def get(self, request, pk, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        dato = CTAS_AHORRO.objects.get(pk=pk)

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="cta_ahos.pdf"'

        # Creamos un objeto PDF con ReportLab
        p = canvas.Canvas(response)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        p.drawString(80, 800, f"Oficina: {dato.oficina}")
        p.drawString(80, 780, f"Línea de Ahorro: {dato.lin_aho}")
        p.drawString(80, 760, f"Nombre Asociado: {dato.asociado}")
        p.drawString(80, 740, f"Número Cuenta: {dato.num_cta}")
        p.drawString(80, 720, f"Estado Cuenta: {dato.est_cta}")
        p.drawString(80, 700, f"Fecha Apertura: {dato.fec_apertura}")
        p.drawString(80, 680, f"Fecha Cancelación: {dato.fec_cancela}")
        p.drawString(80, 660, f"Exenta 4x1000?: {dato.exc_tas_mil}")
        p.drawString(80, 640, f"Fecha Exención: {dato.fec_ini_exc}")

        # Agrega más campos según tus necesidades

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response


class ExportarView(View):
    template_name = 'exportar.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        export_type = request.POST.get('export_type')

        if export_type == 'pdf':
            return self.exportar_pdf()
        elif export_type == 'excel':
            return self.exportar_excel()
        elif export_type == 'csv':
            return self.exportar_csv()

    def exportar_pdf(self):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="exportacion.pdf"'
        cta_ahos = CTAS_AHORRO.objects.all()

        p = canvas.Canvas(response)

        for dato in cta_ahos:
            p.drawString(80, 800, f"Oficina: {dato.oficina}")
            p.drawString(80, 780, f"Línea de Ahorro: {dato.lin_aho}")
            p.drawString(80, 760, f"Nombre Asociado: {dato.asociado}")
            p.drawString(80, 740, f"Número Cuenta: {dato.num_cta}")
            p.drawString(80, 720, f"Estado Cuenta: {dato.est_cta}")
            p.drawString(80, 700, f"Fecha Apertura: {dato.fec_apertura}")
            p.drawString(80, 680, f"Fecha Cancelación: {dato.fec_cancela}")
            p.drawString(80, 660, f"Exenta 4x1000?: {dato.exc_tas_mil}")
            p.drawString(80, 640, f"Fecha Exención: {dato.fec_ini_exc}")

            # Agrega más campos según tus necesidades

            # Agrega un salto de página para el siguiente conjunto de datos
            p.showPage()
        # Puedes agregar más contenido según tus necesidades.
        p.save()

        return response

    def exportar_excel(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportacion.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        # Añade encabezados a la hoja de cálculo utilizando los nombres de campo del modelo
        headers = [field.name for field in CTAS_AHORRO._meta.fields]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Añade datos a la hoja de cálculo
        cta_ahos = CTAS_AHORRO.objects.all()
        for row_num, data in enumerate(cta_ahos, start=2):
            for col_num, field in enumerate(headers, 1):
                sheet.cell(row=row_num, column=col_num,
                           value=getattr(data, field))

        workbook.save(response)
        return response

    def exportar_csv(self):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="exportacion.csv"'

        writer = csv.writer(response)
        # Añade más encabezados según tus necesidades
        writer.writerow(['ID', 'Nombre'])

        cta_ahos = CTAS_AHORRO.objects.all()
        for data in cta_ahos:
            # Añade más campos según tus necesidades
            writer.writerow([data.pk, data.nombre])

        return response
